import openpyxl
from openpyxl import Workbook
wbname = "1.xlsx"
wb = openpyxl.load_workbook(wbname)
sheet = wb.worksheets[0]
print(wb.worksheets)
print(wb.worksheets[0])
print(sheet)
print(sheet.cell(row=2, column=1).value)
sheet.cell(row=2, column=1).value = 5
print(sheet.cell(row=2, column=1).value)
wb.save(wbname)